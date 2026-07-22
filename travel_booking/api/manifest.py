# travel_booking/api/manifest.py
# Aroya Manifest Filler — Cara A (isi template yang Aroya bagi)
#
# Aliran:
#   Admin upload template Aroya (Excel) untuk satu Trip Date
#   -> untuk setiap baris, padan Stateroom # (kol G) + Guest # (kol I)
#      dengan Reservation (res.stateroom_no + res.aroya_guest_no)
#   -> isi kolum traveller (K-Z) dari data Traveller
#   -> export Excel yang dah lengkap + laporan baris tak padan

import frappe
import base64
import io


# ══════════════════════════════════════════════
# KONFIGURASI — senang ubah di sini
# (index kolum, 0-based: A=0, B=1, ... Z=25)
# ══════════════════════════════════════════════

# Kolum kunci untuk padanan
COL_STATEROOM = 6   # G — Stateroom #
COL_GUEST_NO  = 8   # I — Guest #

# Baris data mula (0-based). Row 0 = header, data mula row 1.
DATA_START_ROW = 1

# Mapping: index kolum manifest -> cara dapat nilai dari traveller
# Fungsi lambda terima dict `t` (data traveller) -> pulang nilai untuk sel
COLUMN_FILLERS = {
    10: lambda t: t.get("last_name") or "",                    # K — Last Name
    11: lambda t: t.get("first_name") or "",                   # L — First Name
    12: lambda t: _map_gender(t.get("gender")),                # M — Gender
    13: lambda t: _fmt_date(t.get("date_of_birth")),           # N — Birth Date
    14: lambda t: t.get("nationality") or "",                  # O — Residency
    15: lambda t: t.get("email") or "",                        # P — Email
    16: lambda t: _phone_code(t.get("phone")),                 # Q — Intl.Code
    17: lambda t: t.get("nationality") or "",                  # R — Intl.Code Country
    18: lambda t: _phone_number(t.get("phone")),               # S — Telephone
    19: lambda t: DEFAULT_LANGUAGE,                            # T — Language (default)
    20: lambda t: t.get("nationality") or "",                  # U — Country of Birth
    # V (21) City* — Option B: biar kosong (Traveller tiada)
    # W (22) State — biar kosong
    # X (23) Localized First Name — biar kosong
    # Y (24) Localized Last Name — biar kosong
    # Z (25) SMS Can Contact — biar apa yang Aroya dah isi
}

DEFAULT_LANGUAGE = "English"


# ══════════════════════════════════════════════
# HELPER FORMAT
# ══════════════════════════════════════════════

def _map_gender(gender):
    """Padan format gender Aroya. Laras kalau Aroya guna M/F."""
    if not gender:
        return ""
    return gender  # Male / Female — ubah di sini kalau Aroya nak M/F


def _fmt_date(dob):
    """Format birth date. Laras format ikut kehendak Aroya."""
    if not dob:
        return ""
    try:
        from frappe.utils import getdate
        d = getdate(dob)
        return d.strftime("%Y-%m-%d")   # ubah format di sini kalau perlu
    except Exception:
        return str(dob)


def _phone_code(phone):
    """Ambil kod antarabangsa dari phone. '+60-123456789' -> '60'."""
    if not phone:
        return ""
    p = str(phone).strip()
    if p.startswith("+"):
        rest = p[1:]
        # ambil sebelum '-' atau ' '
        for sep in ["-", " "]:
            if sep in rest:
                return rest.split(sep)[0]
        return rest[:3]
    return ""


def _phone_number(phone):
    """Ambil nombor tanpa kod. '+60-123456789' -> '123456789'."""
    if not phone:
        return ""
    p = str(phone).strip()
    for sep in ["-", " "]:
        if sep in p:
            return p.split(sep, 1)[1].replace(" ", "").replace("-", "")
    if p.startswith("+"):
        return p[3:]
    return p


# ══════════════════════════════════════════════
# BINA PETA RESERVATION UNTUK SAILING
# ══════════════════════════════════════════════

def _build_reservation_map(trip_date):
    """Pulang dict: (stateroom_no, guest_no) -> data traveller.
    Untuk padanan cepat semasa isi manifest.
    """
    rows = frappe.db.sql("""
        SELECT
            res.stateroom_no,
            res.aroya_guest_no,
            t.first_name, t.last_name, t.full_name, t.gender,
            t.date_of_birth, t.nationality, t.email, t.phone,
            t.passport_no, t.passport_expiry
        FROM `tabReservation` res
        JOIN `tabBooking` b ON b.name = res.booking
        LEFT JOIN `tabTraveller` t ON t.name = res.traveller
        WHERE b.trip_date = %s
          AND res.status = 'Confirmed'
          AND res.stateroom_no IS NOT NULL
          AND res.stateroom_no != ''
          AND res.aroya_guest_no IS NOT NULL
    """, trip_date, as_dict=True)

    res_map = {}
    for r in rows:
        key = (str(r.stateroom_no).strip(), int(r.aroya_guest_no))
        res_map[key] = {
            "first_name": r.first_name, "last_name": r.last_name,
            "full_name": r.full_name, "gender": r.gender,
            "date_of_birth": r.date_of_birth, "nationality": r.nationality,
            "email": r.email, "phone": r.phone,
            "passport_no": r.passport_no, "passport_expiry": r.passport_expiry,
            "has_traveller": bool(r.first_name or r.last_name),
        }
    return res_map


# ══════════════════════════════════════════════
# FUNGSI UTAMA — ISI MANIFEST
# ══════════════════════════════════════════════

@frappe.whitelist()
def fill_manifest(trip_date, filedata, filename="manifest.xlsx"):
    """Baca template Aroya, isi data traveller, pulang Excel + laporan.

    Args:
        trip_date: nama Trip Date (sailing)
        filedata: base64 fail template Aroya
        filename: nama fail (untuk output)

    Returns:
        dict: fail (base64) + laporan (filled, unmatched, no_traveller)
    """
    import openpyxl

    # Decode template
    if "," in filedata:
        filedata = filedata.split(",")[1]
    file_bytes = base64.b64decode(filedata)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Bina peta reservation
    res_map = _build_reservation_map(trip_date)

    filled = 0
    unmatched = []       # (stateroom, guest) yang tiada Reservation
    no_traveller = []    # ada Reservation tapi traveller belum diisi

    # Proses setiap baris data
    for row_idx in range(DATA_START_ROW + 1, ws.max_row + 1):  # openpyxl 1-based
        stateroom = ws.cell(row=row_idx, column=COL_STATEROOM + 1).value
        guest_no  = ws.cell(row=row_idx, column=COL_GUEST_NO + 1).value

        if stateroom is None or guest_no is None:
            continue

        key = (str(stateroom).strip(), int(guest_no))
        traveller = res_map.get(key)

        if not traveller:
            unmatched.append({"stateroom": str(stateroom), "guest": int(guest_no)})
            continue

        if not traveller.get("has_traveller"):
            no_traveller.append({"stateroom": str(stateroom), "guest": int(guest_no)})
            continue

        # Isi kolum traveller
        for col_idx, filler in COLUMN_FILLERS.items():
            try:
                value = filler(traveller)
                if value:
                    ws.cell(row=row_idx, column=col_idx + 1).value = value
            except Exception:
                pass

        filled += 1

    # Simpan ke bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    out_b64 = base64.b64encode(out.read()).decode()

    return {
        "status": "ok",
        "filename": "FILLED-" + filename,
        "filedata": out_b64,
        "report": {
            "filled": filled,
            "unmatched_count": len(unmatched),
            "unmatched": unmatched[:50],
            "no_traveller_count": len(no_traveller),
            "no_traveller": no_traveller[:50],
        },
    }


@frappe.whitelist()
def preview_manifest_match(trip_date):
    """Preview — tunjuk padanan sedia ada untuk trip date, tanpa fail.
    Berguna untuk admin semak sebelum export.
    """
    res_map = _build_reservation_map(trip_date)
    matches = []
    for (stateroom, guest), t in sorted(res_map.items()):
        matches.append({
            "stateroom": stateroom,
            "guest": guest,
            "traveller": t.get("full_name") or "(belum diisi)",
            "has_traveller": t.get("has_traveller"),
        })
    return {"status": "ok", "total": len(matches), "matches": matches}